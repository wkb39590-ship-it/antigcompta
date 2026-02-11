





# import os
# from datetime import date
# from pathlib import Path
# from typing import Optional

# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment

# from reportlab.lib.pagesizes import A4, landscape
# from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
# from reportlab.lib import colors
# from reportlab.lib.styles import getSampleStyleSheet

# DED_HEADERS = [
#     "FACT_NUM", "DESIGNATION", "M_HT", "TVA", "M_TTC",
#     "IF", "LIB_FRSS", "ICE_FRS", "TAUX", "DATE_PAI", "DATE_FAC"
# ]

# OUT_DIR = os.getenv("DED_OUT_DIR", "/app/generated/ded")

# def _safe_taux(m_ht: Optional[float], tva: Optional[float], taux: Optional[float]) -> float:
#     if taux is not None:
#         return float(taux)
#     if not m_ht or not tva or float(m_ht) == 0:
#         return 0.0
#     return round((float(tva) / float(m_ht)) * 100, 2)

# def _ded_row(facture) -> list:
#     taux = _safe_taux(facture.montant_ht, facture.montant_tva, getattr(facture, "taux_tva", None))
#     date_paie = getattr(facture, "date_paie", None) or date.today()

#     return [
#         getattr(facture, "numero_facture", None) or "",
#         getattr(facture, "designation", None) or "",
#         float(getattr(facture, "montant_ht", 0) or 0),
#         float(getattr(facture, "montant_tva", 0) or 0),
#         float(getattr(facture, "montant_ttc", 0) or 0),
#         getattr(facture, "if_frs", None) or "",
#         getattr(facture, "fournisseur", None) or "",
#         getattr(facture, "ice_frs", None) or "",
#         taux,
#         date_paie.strftime("%Y-%m-%d"),
#         (facture.date_facture.strftime("%Y-%m-%d") if getattr(facture, "date_facture", None) else ""),
#     ]

# def generate_ded_xlsx(facture) -> str:
#     Path(OUT_DIR).mkdir(parents=True, exist_ok=True)
#     path = os.path.join(OUT_DIR, f"ded_facture_{facture.id}.xlsx")

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "DED"

#     ws.append(DED_HEADERS)
#     for col in range(1, len(DED_HEADERS) + 1):
#         c = ws.cell(row=1, column=col)
#         c.font = Font(bold=True)
#         c.alignment = Alignment(horizontal="center")

#     ws.append(_ded_row(facture))

#     # Widths
#     for i, h in enumerate(DED_HEADERS, start=1):
#         ws.column_dimensions[chr(64 + i)].width = max(12, len(h) + 2)

#     wb.save(path)
#     return path

# def generate_ded_pdf(facture) -> str:
#     Path(OUT_DIR).mkdir(parents=True, exist_ok=True)
#     path = os.path.join(OUT_DIR, f"ded_facture_{facture.id}.pdf")

#     doc = SimpleDocTemplate(
#         path,
#         pagesize=landscape(A4),
#         rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24
#     )
#     styles = getSampleStyleSheet()

#     data = [DED_HEADERS, _ded_row(facture)]
#     table = Table(data)

#     table.setStyle(TableStyle([
#         ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
#         ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
#         ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
#         ("ALIGN", (0, 0), (-1, -1), "CENTER"),
#         ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
#         ("BACKGROUND", (0, 1), (-1, 1), colors.whitesmoke),
#         ("FONTSIZE", (0, 0), (-1, -1), 9),
#         ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
#     ]))

#     story = [
#         Paragraph("Relevé de déduction TVA (DGI) - Généré automatiquement", styles["Title"]),
#         Spacer(1, 12),
#         table
#     ]
#     doc.build(story)
#     return path
















import os
from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

DED_HEADERS = [
    "FACT_NUM", "DESIGNATION", "M_HT", "TVA", "M_TTC",
    "IF", "LIB_FRSS", "ICE_FRS", "TAUX", "DATE_PAI", "DATE_FAC"
]

OUT_DIR = os.getenv("DED_OUT_DIR", "/app/generated/ded")


def _safe_taux(m_ht: Optional[float], tva: Optional[float], taux: Optional[float]) -> float:
    if taux is not None:
        return float(taux)
    if not m_ht or not tva or float(m_ht) == 0:
        return 0.0
    return round((float(tva) / float(m_ht)) * 100, 2)


def _ded_row(facture) -> list:
    taux = _safe_taux(facture.montant_ht, facture.montant_tva, getattr(facture, "taux_tva", None))
    date_paie = getattr(facture, "date_paie", None) or date.today()

    return [
        getattr(facture, "numero_facture", None) or "",
        getattr(facture, "designation", None) or "",
        float(getattr(facture, "montant_ht", 0) or 0),
        float(getattr(facture, "montant_tva", 0) or 0),
        float(getattr(facture, "montant_ttc", 0) or 0),
        getattr(facture, "if_frs", None) or "",
        getattr(facture, "fournisseur", None) or "",
        getattr(facture, "ice_frs", None) or "",
        float(taux),
        date_paie.strftime("%Y-%m-%d"),
        (facture.date_facture.strftime("%Y-%m-%d") if getattr(facture, "date_facture", None) else ""),
    ]


def generate_ded_xlsx(facture) -> str:
    Path(OUT_DIR).mkdir(parents=True, exist_ok=True)
    path = os.path.join(OUT_DIR, f"ded_facture_{facture.id}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "DED"

    ws.append(DED_HEADERS)
    for col in range(1, len(DED_HEADERS) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    ws.append(_ded_row(facture))

    # Widths (simple)
    for i, h in enumerate(DED_HEADERS, start=1):
        # NOTE: chr(64+i) marche jusqu'à Z; ici on a 11 colonnes donc OK.
        ws.column_dimensions[chr(64 + i)].width = max(12, len(h) + 2)

    wb.save(path)
    return path


def generate_ded_pdf(facture) -> str:
    """
    Version corrigée:
    - A4 paysage
    - colWidths fixes
    - wrap sur DESIGNATION et LIB_FRSS
    - police réduite + padding
    """
    Path(OUT_DIR).mkdir(parents=True, exist_ok=True)
    path = os.path.join(OUT_DIR, f"ded_facture_{facture.id}.pdf")

    doc = SimpleDocTemplate(
        path,
        pagesize=landscape(A4),
        leftMargin=12,
        rightMargin=12,
        topMargin=12,
        bottomMargin=12,
    )

    styles = getSampleStyleSheet()

    # Style pour wrapper texte long
    wrap_style = ParagraphStyle(
        name="wrap7",
        parent=styles["Normal"],
        fontSize=7,
        leading=8,
        spaceBefore=0,
        spaceAfter=0,
    )

    def W(x):
        return Paragraph(str(x or ""), wrap_style)

    row = _ded_row(facture)

    # Wrap seulement sur colonnes longues
    row_wrapped = [
        row[0],          # FACT_NUM
        W(row[1]),       # DESIGNATION (wrap)
        f"{row[2]:.2f}", # M_HT
        f"{row[3]:.2f}", # TVA
        f"{row[4]:.2f}", # M_TTC
        row[5],          # IF
        W(row[6]),       # LIB_FRSS (wrap)
        row[7],          # ICE_FRS
        f"{row[8]:.2f}", # TAUX
        row[9],          # DATE_PAI
        row[10],         # DATE_FAC
    ]

    data = [DED_HEADERS, row_wrapped]

    # Largeurs colonne (adaptées A4 paysage)
    # total ~ 842 - 24 = 818 points dispo, ces widths passent bien
    col_widths = [60, 185, 52, 45, 55, 65, 190, 90, 40, 65, 65]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),

        # Body
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("LEADING", (0, 1), (-1, -1), 8),
        ("VALIGN", (0, 1), (-1, -1), "TOP"),

        # Align numérique au centre (ou RIGHT si tu préfères)
        ("ALIGN", (2, 1), (4, 1), "CENTER"),  # montants
        ("ALIGN", (8, 1), (8, 1), "CENTER"),  # taux
        ("ALIGN", (0, 1), (0, 1), "CENTER"),  # fact_num

        # Grid + padding
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),

        # Background body
        ("BACKGROUND", (0, 1), (-1, 1), colors.whitesmoke),
    ]))

    story = [
        Paragraph("Relevé de déduction TVA (DGI) - Généré automatiquement", styles["Title"]),
        Spacer(1, 10),
        table,
    ]

    doc.build(story)
    return path
