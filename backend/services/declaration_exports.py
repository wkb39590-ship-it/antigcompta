# services/declaration_exports.py
from io import BytesIO
from datetime import datetime
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


HEADERS = [
    "FACT_NUM", "DESIGNATION", "M_HT", "TVA", "M_TTC", "IF",
    "LIB_FRSS", "ICE_FRS", "TAUX", "DATE_PAI", "DATE_FAC"
]


def build_excel(declaration_dict: Dict[str, Any]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{declaration_dict['annee']}-{declaration_dict['mois']:02d}"

    title = f"Déclaration Mensuelle DGI - {declaration_dict['annee']}-{declaration_dict['mois']:02d}"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    # Header row
    header_row = 3
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4F81BD")
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    row = header_row + 1
    for ln in declaration_dict["lignes"]:
        values = [
            ln.get("fact_num"),
            ln.get("designation"),
            ln.get("m_ht"),
            ln.get("tva"),
            ln.get("m_ttc"),
            ln.get("if_frs"),
            ln.get("lib_frss"),
            ln.get("ice_frs"),
            ln.get("taux"),
            str(ln.get("date_pai")) if ln.get("date_pai") else None,
            str(ln.get("date_fac")) if ln.get("date_fac") else None,
        ]
        for col, v in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=v)
        row += 1

    # Totals
    ws.cell(row=row + 1, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row + 1, column=3, value=declaration_dict["total_ht"]).font = Font(bold=True)
    ws.cell(row=row + 1, column=4, value=declaration_dict["total_tva"]).font = Font(bold=True)
    ws.cell(row=row + 1, column=5, value=declaration_dict["total_ttc"]).font = Font(bold=True)

    # Column widths
    widths = [14, 25, 12, 12, 12, 12, 28, 18, 10, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def build_pdf(declaration_dict: Dict[str, Any]) -> BytesIO:
    out = BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()

    story = []
    title = f"Déclaration Mensuelle DGI - {declaration_dict['annee']}-{declaration_dict['mois']:02d}"
    story.append(Paragraph(title, styles["Title"]))
    story.append(Spacer(1, 10))

    data = [HEADERS]
    for ln in declaration_dict["lignes"]:
        data.append([
            ln.get("fact_num") or "",
            ln.get("designation") or "",
            f"{ln.get('m_ht') or 0:.2f}" if ln.get("m_ht") is not None else "",
            f"{ln.get('tva') or 0:.2f}" if ln.get("tva") is not None else "",
            f"{ln.get('m_ttc') or 0:.2f}" if ln.get("m_ttc") is not None else "",
            ln.get("if_frs") or "",
            ln.get("lib_frss") or "",
            ln.get("ice_frs") or "",
            f"{ln.get('taux') or 0:.2f}" if ln.get("taux") is not None else "",
            str(ln.get("date_pai")) if ln.get("date_pai") else "",
            str(ln.get("date_fac")) if ln.get("date_fac") else "",
        ])

    # Totals row
    data.append([
        "TOTAL", "", f"{declaration_dict['total_ht']:.2f}", f"{declaration_dict['total_tva']:.2f}",
        f"{declaration_dict['total_ttc']:.2f}", "", "", "", "", "", ""
    ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 1), (-1, -2), colors.whitesmoke),
        ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))

    story.append(table)
    doc.build(story)

    out.seek(0)
    return out
