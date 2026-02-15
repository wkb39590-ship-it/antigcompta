# services/exports.py
from pathlib import Path
from typing import List

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADERS = ["FACT_NUM", "DESIGNATION", "M_HT", "TVA", "M_TTC", "IF", "LIB_FRSS", "ICE_FRS", "TAUX", "DATE_PAI", "DATE_FAC"]


def export_declaration_pdf(decl, lignes: List, out_path: str) -> str:
    out_path = str(out_path)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        out_path,
        pagesize=landscape(A4),
        leftMargin=18,
        rightMargin=18,
        topMargin=18,
        bottomMargin=18,
        title=f"Declaration DGI {decl.annee}-{decl.mois:02d}",
    )

    styles = getSampleStyleSheet()
    title = Paragraph(f"<b>Déclaration Mensuelle DGI</b> — {decl.annee}-{decl.mois:02d}", styles["Title"])
    subtitle = Paragraph("Relevé de déduction (format interne)", styles["Normal"])

    data = [HEADERS]
    for l in lignes:
        data.append([
            l.fact_num or "",
            l.designation or "",
            f"{float(l.m_ht or 0.0):.2f}",
            f"{float(l.tva or 0.0):.2f}",
            f"{float(l.m_ttc or 0.0):.2f}",
            l.if_frs or "",
            l.lib_frss or "",
            l.ice_frs or "",
            f"{float(l.taux or 0.0):.2f}",
            str(l.date_pai) if l.date_pai else "",
            str(l.date_fac) if l.date_fac else "",
        ])

    # total row
    data.append([
        "TOTAL", "",
        f"{float(decl.total_ht or 0.0):.2f}",
        f"{float(decl.total_tva or 0.0):.2f}",
        f"{float(decl.total_ttc or 0.0):.2f}",
        "", "", "", "", "", ""
    ])

    # Column widths (landscape A4)
    col_widths = [2.6*cm, 4.0*cm, 2.2*cm, 2.2*cm, 2.4*cm, 2.4*cm, 6.0*cm, 5.0*cm, 2.0*cm, 2.8*cm, 2.8*cm]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    header_bg = colors.HexColor("#1E3A8A")   # bleu moderne
    header_fg = colors.white
    zebra_a = colors.whitesmoke
    zebra_b = colors.HexColor("#F3F4F6")

    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), header_bg),
        ("TEXTCOLOR", (0, 0), (-1, 0), header_fg),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (2, 1), (4, -1), "RIGHT"),
        ("ALIGN", (8, 1), (8, -1), "RIGHT"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [zebra_a, zebra_b]),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E5E7EB")),
        ("LINEABOVE", (0, -1), (-1, -1), 1.0, colors.HexColor("#111827")),
    ])

    table.setStyle(style)

    story = [title, Spacer(1, 6), subtitle, Spacer(1, 14), table]
    doc.build(story)
    return out_path


def export_declaration_excel(decl, lignes: List, out_path: str) -> str:
    out_path = str(out_path)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{decl.annee}-{decl.mois:02d}"

    # Styles
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(HEADERS)

    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for l in lignes:
        ws.append([
            l.fact_num or "",
            l.designation or "",
            float(l.m_ht or 0.0),
            float(l.tva or 0.0),
            float(l.m_ttc or 0.0),
            l.if_frs or "",
            l.lib_frss or "",
            l.ice_frs or "",
            float(l.taux or 0.0),
            str(l.date_pai) if l.date_pai else "",
            str(l.date_fac) if l.date_fac else "",
        ])

    # Total row
    ws.append([
        "TOTAL", "",
        float(decl.total_ht or 0.0),
        float(decl.total_tva or 0.0),
        float(decl.total_ttc or 0.0),
        "", "", "", "", "", ""
    ])

    last_row = ws.max_row

    # Format & borders
    for r in range(2, last_row + 1):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c in (3, 4, 5, 9):
                cell.number_format = "0.00"
                cell.alignment = right
            elif c in (1, 10, 11):
                cell.alignment = center
            elif c == 2 or c == 7:
                cell.alignment = left
            else:
                cell.alignment = center

    # Total row styling
    total_fill = PatternFill("solid", fgColor="E5E7EB")
    total_font = Font(bold=True)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=last_row, column=c)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border

    # Auto width
    for c in range(1, len(HEADERS) + 1):
        col = get_column_letter(c)
        max_len = 0
        for r in range(1, last_row + 1):
            v = ws.cell(row=r, column=c).value
            max_len = max(max_len, len(str(v)) if v is not None else 0)
        ws.column_dimensions[col].width = min(max(10, max_len + 2), 45)

    ws.freeze_panes = "A2"
    wb.save(out_path)
    return out_path
